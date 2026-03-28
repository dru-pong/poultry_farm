from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal
import re


# ═════════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ═════════════════════════════════════════════════════════════════════════

def _adjust_column_widths(ws):
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


HEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=16, color='1F4E79')
SECTION_FONT = Font(bold=True, size=14)
GREEN_FONT = Font(bold=True, size=14, color='006400')
RED_FONT = Font(bold=True, size=14, color='8B0000')


def _write_headers(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL


# ═════════════════════════════════════════════════════════════════════════
# EXPENSES EXCEL
# ═════════════════════════════════════════════════════════════════════════

def create_expenses_excel_report(report_data, filename='expenses_report.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet('Summary')
    _create_summary_sheet(summary_ws, report_data)

    category_ws = wb.create_sheet('By Category')
    _create_category_sheet(category_ws, report_data)

    daily_ws = wb.create_sheet('Daily Expenses')
    _create_daily_sheet(daily_ws, report_data)

    detail_ws = wb.create_sheet('All Transactions')
    _create_detail_sheet(detail_ws, report_data)

    for ws in wb.worksheets:
        _adjust_column_widths(ws)
    return wb


def _create_summary_sheet(ws, data):
    ws['A1'] = 'POULTRY FARM - EXPENSE REPORT'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = 'Report Period:'
    ws['B3'] = f"{data['period']['start_date']} to {data['period']['end_date']}"
    ws['A5'] = 'Metric'
    ws['B5'] = 'Value'
    metrics = [
        ('Total Expenses', data['total_expenses']),
        ('Average per Day', data['average_expense_per_day']),
        ('Number of Days', data['number_of_days']),
        ('Total Transactions', data['expense_count']),
    ]
    for i, (label, value) in enumerate(metrics, start=6):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = f'₵{value}' if 'Expense' in label or 'per Day' in label else value
    for cell in ['A5', 'B5']:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL


def _create_category_sheet(ws, data):
    ws['A1'] = 'EXPENSES BY CATEGORY'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Category', 'Total Amount', 'Transaction Count', 'Percentage'])
    total = float(data['total_expenses'])
    for i, cat in enumerate(data['category_breakdown'], start=4):
        ws.cell(row=i, column=1, value=cat['category'])
        ws.cell(row=i, column=2, value=f"₵{cat['total']}")
        ws.cell(row=i, column=3, value=cat['count'])
        pct = (float(cat['total']) / total * 100) if total > 0 else 0
        ws.cell(row=i, column=4, value=f"{pct:.1f}%")


def _create_daily_sheet(ws, data):
    ws['A1'] = 'DAILY EXPENSES'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Date', 'Amount'])
    for i, (d, amt) in enumerate(data['daily_expenses'].items(), start=4):
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=2, value=f"₵{amt}")


def _create_detail_sheet(ws, data):
    ws['A1'] = 'ALL EXPENSE TRANSACTIONS'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Date', 'Category', 'Description', 'Amount', 'Notes'])
    row = 4
    for cat in data['category_breakdown']:
        for item in cat['items']:
            ws.cell(row=row, column=1, value=item['date'])
            ws.cell(row=row, column=2, value=cat['category'])
            ws.cell(row=row, column=3, value=item['description'])
            ws.cell(row=row, column=4, value=f"₵{item['amount']}")
            ws.cell(row=row, column=5, value=item['notes'])
            row += 1


# ═════════════════════════════════════════════════════════════════════════
# SALES EXCEL
# ═════════════════════════════════════════════════════════════════════════

def create_sales_excel_report(report_data, filename='sales_report.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet('Summary')
    _create_sales_summary_sheet(summary_ws, report_data)

    customer_ws = wb.create_sheet('By Customer')
    _create_sales_by_customer_sheet(customer_ws, report_data)

    daily_ws = wb.create_sheet('Daily Sales')
    _create_sales_daily_sheet(daily_ws, report_data)

    detail_ws = wb.create_sheet('All Transactions')
    _create_sales_detail_sheet(detail_ws, report_data)

    eggtype_ws = wb.create_sheet('By Egg Type')
    _create_sales_by_eggtype_sheet(eggtype_ws, report_data)

    for ws in wb.worksheets:
        _adjust_column_widths(ws)
    return wb


def _create_sales_summary_sheet(ws, data):
    ws['A1'] = 'POULTRY FARM - SALES REPORT'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = 'Report Period:'
    ws['B3'] = f"{data['period']['start_date']} to {data['period']['end_date']}"
    ws['A5'] = 'Metric'
    ws['B5'] = 'Value'
    metrics = [
        ('Total Sales', data['total_sales']),
        ('Total Revenue', f"₵{data['total_revenue']}"),
        ('Retail Transactions', data['retail_count']),
        ('Retail Revenue', f"₵{data['retail_revenue']}"),
        ('Wholesale Transactions', data['wholesale_count']),
        ('Wholesale Revenue', f"₵{data['wholesale_revenue']}"),
    ]
    for i, (label, value) in enumerate(metrics, start=6):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    for cell in ['A5', 'B5']:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    ws['A15'] = 'Quantity Breakdown (Crates)'
    ws['A15'].font = SECTION_FONT
    ws['A17'] = 'Egg Type'
    ws['B17'] = 'Quantity'
    quantities = data.get('quantities', {})
    for i, (et, qty) in enumerate([
        ('Broken', quantities.get('broken', 0)),
        ('Small', quantities.get('small', 0)),
        ('Medium', quantities.get('medium', 0)),
        ('Big', quantities.get('big', 0))
    ], start=18):
        ws[f'A{i}'] = et
        ws[f'B{i}'] = qty
        ws[f'B{i}'].font = HEADER_FONT


def _create_sales_by_customer_sheet(ws, data):
    ws['A1'] = 'SALES BY CUSTOMER'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Customer', 'Total Purchases', 'Transaction Count', 'Avg per Transaction', 'Percentage'])
    total = float(data['total_revenue'])
    for i, c in enumerate(data.get('customer_breakdown', []), start=4):
        ws.cell(row=i, column=1, value=c['customer'])
        ws.cell(row=i, column=2, value=f"₵{c['total']}")
        ws.cell(row=i, column=3, value=c['transaction_count'])
        avg = float(c['total']) / c['transaction_count'] if c['transaction_count'] > 0 else 0
        ws.cell(row=i, column=4, value=f"₵{avg:.2f}")
        pct = (float(c['total']) / total * 100) if total > 0 else 0
        ws.cell(row=i, column=5, value=f"{pct:.1f}%")


def _create_sales_daily_sheet(ws, data):
    ws['A1'] = 'DAILY SALES'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Date', 'Sales Count', 'Revenue'])
    for i, (d, sd) in enumerate(data.get('daily_sales', {}).items(), start=4):
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=2, value=sd.get('count', 0))
        ws.cell(row=i, column=3, value=f"₵{sd.get('revenue', '0.00')}")


def _create_sales_detail_sheet(ws, data):
    ws['A1'] = 'ALL SALES TRANSACTIONS'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Date', 'Time', 'Customer', 'Sale Type', 'Egg Type', 'Quantity', 'Price/Crate', 'Line Total', 'Notes'])
    row = 4
    for t in data.get('transactions', []):
        dt = t.get('sale_datetime', '')
        date_part = dt.split('T')[0] if 'T' in dt else dt.split(' ')[0]
        time_part = dt.split('T')[1][:5] if 'T' in dt else ''
        ws.cell(row=row, column=1, value=date_part)
        ws.cell(row=row, column=2, value=time_part)
        ws.cell(row=row, column=3, value=t.get('customer', 'Retail'))
        ws.cell(row=row, column=4, value=t.get('sale_type', '').title())
        ws.cell(row=row, column=5, value=t.get('egg_type', ''))
        ws.cell(row=row, column=6, value=t.get('quantity', 0))
        ws.cell(row=row, column=7, value=f"₵{t.get('price_per_crate', '0.00')}")
        ws.cell(row=row, column=8, value=f"₵{t.get('line_total', '0.00')}")
        ws.cell(row=row, column=9, value=t.get('notes', ''))
        row += 1


def _create_sales_by_eggtype_sheet(ws, data):
    ws['A1'] = 'SALES BY EGG TYPE'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Egg Type', 'Total Crates Sold', 'Revenue', 'Percentage of Total'])
    total_revenue = float(data['total_revenue'])
    for i, et in enumerate(data.get('egg_type_breakdown', []), start=4):
        ws.cell(row=i, column=1, value=et['egg_type'])
        ws.cell(row=i, column=2, value=et['total_crates'])
        ws.cell(row=i, column=3, value=f"₵{et['revenue']}")
        pct = (float(et['revenue']) / total_revenue * 100) if total_revenue > 0 else 0
        ws.cell(row=i, column=4, value=f"{pct:.1f}%")


# ═════════════════════════════════════════════════════════════════════════
# PROFIT & LOSS EXCEL
# ═════════════════════════════════════════════════════════════════════════

def create_profit_loss_excel_report(report_data, filename='profit_loss_report.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet('P&L Summary')
    _create_pl_summary_sheet(summary_ws, report_data)

    revenue_ws = wb.create_sheet('Revenue Breakdown')
    _create_pl_revenue_sheet(revenue_ws, report_data)

    expense_ws = wb.create_sheet('Expense Breakdown')
    _create_pl_expense_sheet(expense_ws, report_data)

    for ws in wb.worksheets:
        _adjust_column_widths(ws)
    return wb


def _create_pl_summary_sheet(ws, data):
    ws['A1'] = 'POULTRY FARM - PROFIT & LOSS STATEMENT'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = 'Report Period:'
    ws['B3'] = f"{data['period']['start_date']} to {data['period']['end_date']}"
    ws['A5'] = 'REVENUE'
    ws['A5'].font = GREEN_FONT
    ws['A7'] = 'Retail Sales'
    ws['B7'] = f"₵{data.get('revenue_breakdown', {}).get('retail', '0.00')}"
    ws['A8'] = 'Wholesale Sales'
    ws['B8'] = f"₵{data.get('revenue_breakdown', {}).get('wholesale', '0.00')}"
    ws['A10'] = 'TOTAL REVENUE'
    ws['B10'] = f"₵{data.get('total_revenue', '0.00')}"
    ws['B10'].font = HEADER_FONT
    ws['A12'] = 'EXPENSES'
    ws['A12'].font = RED_FONT
    ws['A14'] = 'Total Expenses'
    ws['B14'] = f"₵{data.get('total_expenses', '0.00')}"
    ws['A17'] = 'NET PROFIT'
    ws['A17'].font = SECTION_FONT
    net = float(data.get('net_profit', 0))
    ws['B17'] = f"₵{data.get('net_profit', '0.00')}"
    ws['B17'].font = Font(bold=True, size=14, color='006400' if net >= 0 else '8B0000')
    ws['A18'] = 'Profit Margin'
    ws['B18'] = f"{data.get('profit_margin', 0)}%"
    for cell in ['A5', 'A12', 'A17']:
        ws[cell].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')


def _create_pl_revenue_sheet(ws, data):
    ws['A1'] = 'REVENUE BREAKDOWN'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Sale Type', 'Transaction Count', 'Total Revenue', 'Percentage'])
    total = float(data.get('total_revenue', 0))
    rev = data.get('revenue_breakdown', {})
    for i, (label, count_key, rev_key) in enumerate([
        ('Retail', 'retail_count', 'retail'),
        ('Wholesale', 'wholesale_count', 'wholesale'),
    ], start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=data.get(count_key, 0))
        ws.cell(row=i, column=3, value=f"₵{rev.get(rev_key, '0.00')}")
        pct = (float(rev.get(rev_key, 0)) / total * 100) if total > 0 else 0
        ws.cell(row=i, column=4, value=f"{pct:.1f}%")


def _create_pl_expense_sheet(ws, data):
    ws['A1'] = 'EXPENSE BREAKDOWN'
    ws['A1'].font = SECTION_FONT
    _write_headers(ws, 3, ['Category', 'Total Amount', 'Percentage'])
    total = float(data.get('total_expenses', 0))
    for i, exp in enumerate(data.get('expense_breakdown', []), start=4):
        cat = exp.get('category__name', 'Unknown')
        amt = exp.get('total', 0)
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=f"₵{amt}")
        pct = (float(amt) / total * 100) if total > 0 else 0
        ws.cell(row=i, column=3, value=f"{pct:.1f}%")


# ═════════════════════════════════════════════════════════════════════════
# CUSTOMER REPORT EXCEL  (NEW)
# ═════════════════════════════════════════════════════════════════════════

def _safe_sheet_name(name, max_len=31):
    """Sanitise a string for use as an Excel sheet name."""
    cleaned = re.sub(r'[\[\]:*?/\\]', '', name)
    return cleaned[:max_len] if cleaned else 'Customer'


def create_customer_excel_report(report_data):
    """
    Creates a workbook with:
      1. Summary sheet — all customers ranked by outstanding balance
      2. One sheet per customer — egg breakdown, daily history, credit payments
    """
    wb = Workbook()
    wb.remove(wb.active)

    period = report_data['period']
    customers = report_data['customers']

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws = wb.create_sheet('Summary')
    ws['A1'] = 'POULTRY FARM - CUSTOMER REPORT'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = 'Report Period:'
    ws['B3'] = f"{period['start_date']} to {period['end_date']}"
    ws['A4'] = 'Customers with activity:'
    ws['B4'] = report_data['customer_count']

    _write_headers(ws, 6, [
        'Customer', 'Phone', 'Transactions',
        'Total Purchased', 'Total Paid (range)', 'Outstanding Balance'
    ])

    for i, c in enumerate(customers, start=7):
        s = c['summary']
        ws.cell(row=i, column=1, value=c['customer_name'])
        ws.cell(row=i, column=2, value=c.get('phone', ''))
        ws.cell(row=i, column=3, value=s['transaction_count'])
        ws.cell(row=i, column=4, value=f"₵{s['total_purchased_in_range']}")
        ws.cell(row=i, column=5, value=f"₵{s['total_paid_in_range']}")
        ws.cell(row=i, column=6, value=f"₵{s['outstanding_balance_alltime']}")

        # Highlight outstanding > 0
        bal = Decimal(s['outstanding_balance_alltime'])
        if bal > 0:
            ws.cell(row=i, column=6).font = Font(bold=True, color='8B0000')

    _adjust_column_widths(ws)

    # ── Per-customer sheets ─────────────────────────────────────────────
    seen_names = {}
    for c in customers:
        base = _safe_sheet_name(c['customer_name'])
        # Handle duplicate names
        if base in seen_names:
            seen_names[base] += 1
            sheet_name = f"{base[:28]}_{seen_names[base]}"
        else:
            seen_names[base] = 1
            sheet_name = base

        cws = wb.create_sheet(sheet_name)
        _write_customer_sheet(cws, c, period)
        _adjust_column_widths(cws)

    return wb


def _write_customer_sheet(ws, customer_data, period):
    """Write a single customer's detail sheet."""
    c = customer_data
    s = c['summary']

    # Title
    ws['A1'] = c['customer_name']
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Period: {period['start_date']} to {period['end_date']}"
    ws['A2'].font = Font(italic=True, color='666666')

    # ── Summary section ─────────────────────────────────────────────────
    row = 4
    ws.cell(row=row, column=1, value='SUMMARY').font = SECTION_FONT
    row += 1
    for label, val in [
        ('Total Purchased (range)', f"₵{s['total_purchased_in_range']}"),
        ('Total Paid (range)', f"₵{s['total_paid_in_range']}"),
        ('Transactions', s['transaction_count']),
        ('Outstanding Balance (all-time)', f"₵{s['outstanding_balance_alltime']}"),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1

    bal = Decimal(s['outstanding_balance_alltime'])
    if bal > 0:
        ws.cell(row=row - 1, column=2).font = Font(bold=True, color='8B0000')

    # ── Egg type breakdown ──────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value='EGG TYPE BREAKDOWN').font = SECTION_FONT
    row += 1
    _write_headers(ws, row, ['Egg Type', 'Crates', 'Revenue'])
    row += 1
    for eb in c.get('egg_breakdown', []):
        ws.cell(row=row, column=1, value=eb['egg_type'])
        ws.cell(row=row, column=2, value=eb['crates'])
        ws.cell(row=row, column=3, value=f"₵{eb['revenue']}")
        row += 1

    # ── Daily purchase history ──────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value='PURCHASE HISTORY (active days only)').font = SECTION_FONT
    row += 1
    _write_headers(ws, row, [
        'Date', 'Egg Type', 'Crates', 'Price/Crate',
        'Sale Total', 'Amount Paid', 'Status', 'Running Balance'
    ])
    row += 1
    for entry in c.get('daily_history', []):
        items = entry.get('items', [])
        if not items:
            # Sale with no item detail (shouldn't happen but be safe)
            ws.cell(row=row, column=1, value=entry['date'])
            ws.cell(row=row, column=5, value=f"₵{entry['sale_total']}")
            ws.cell(row=row, column=6, value=f"₵{entry['amount_paid']}")
            ws.cell(row=row, column=7, value=entry['payment_status'])
            ws.cell(row=row, column=8, value=f"₵{entry['running_balance']}")
            row += 1
        else:
            for idx, item in enumerate(items):
                ws.cell(row=row, column=1, value=entry['date'] if idx == 0 else '')
                ws.cell(row=row, column=2, value=item['egg_type'])
                ws.cell(row=row, column=3, value=item['crates'])
                ws.cell(row=row, column=4, value=f"₵{item['price_per_crate']}")
                # Only show totals on first item row
                if idx == 0:
                    ws.cell(row=row, column=5, value=f"₵{entry['sale_total']}")
                    ws.cell(row=row, column=6, value=f"₵{entry['amount_paid']}")
                    ws.cell(row=row, column=7, value=entry['payment_status'])
                    ws.cell(row=row, column=8, value=f"₵{entry['running_balance']}")
                row += 1

    # ── Credit payment history ──────────────────────────────────────────
    credit_payments = c.get('credit_payments', [])
    if credit_payments:
        row += 1
        ws.cell(row=row, column=1, value='CREDIT PAYMENTS').font = SECTION_FONT
        row += 1
        _write_headers(ws, row, ['Date', 'Amount', 'Against Sale #', 'Notes'])
        row += 1
        for cp in credit_payments:
            ws.cell(row=row, column=1, value=cp['date'])
            ws.cell(row=row, column=2, value=f"₵{cp['amount']}")
            ws.cell(row=row, column=3, value=cp.get('sale_id') or 'General')
            ws.cell(row=row, column=4, value=cp.get('notes', ''))
            row += 1